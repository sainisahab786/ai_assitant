import os
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import imagehash
import cv2
import pandas as pd
import re
import base64


# Load configuration from config.json
with open('config.json', 'r') as config_file:
    config = json.load(config_file)

main_folder_path = config['main_folder_path']
comparison_images_folder_path = config['comparison_images_folder_path']
output_path = config['output_path']
images_output_path = config['images_output_path']
base_url = config['base_url']
image_similarity_threshold = config['image_similarity_threshold']

# Create output directories if they don't exist
for path in [output_path, images_output_path]:
    if not os.path.exists(path):
        os.makedirs(path)

# Function to calculate perceptual hash for an image
def get_image_hash(image_path):
    img = PILImage.open(image_path)
    return imagehash.phash(img)  # Using perceptual hash (phash)

# Helper function to check if an image is similar (within a threshold)
def is_similar_image(image_file, comparison_hashes, threshold=5):
    image_hash = get_image_hash(image_file)
    for comparison_hash in comparison_hashes:
        if abs(image_hash - comparison_hash) <= threshold:
            return True  # Images are similar if difference is within threshold
    return False

# Function to detect barcodes using OpenCV
def contains_barcode(image_path):
    img = cv2.imread(image_path)
    gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    barcode_detector = cv2.QRCodeDetector()
    data, _, _ = barcode_detector.detectAndDecode(gray_img)
    return bool(data)

# Load and hash all images from the comparison images folder
def load_comparison_image_hashes(folder_path):
    comparison_hashes = []
    for filename in os.listdir(folder_path):
        if filename.endswith((".png", ".jpg", ".jpeg")):
            image_path = os.path.join(folder_path, filename)
            comparison_hashes.append(get_image_hash(image_path))
    return comparison_hashes

# Regular expression pattern for report numbers
report_number_pattern = re.compile(r"(Report Number[s]*)\s+([A-Za-z0-9]+)")

# Function to extract report numbers into two columns
def extract_report_numbers(df):
    for idx, row in df.iterrows():
        for col_idx, cell_value in enumerate(row):
            if isinstance(cell_value, str):
                match = report_number_pattern.search(cell_value)
                if match:
                    df.at[idx, col_idx] = match.group(1)  # "Report Number"
                    df.at[idx, col_idx + 1] = match.group(2)  # Actual number
    return df

# Function to get report number from DataFrame
def get_report_number_from_df(df):
    """Extract report number from DataFrame"""
    for idx, row in df.iterrows():
        for col in df.columns:
            value = row[col]
            if isinstance(value, str):
                # Check for "Report Number" format
                match = re.search(r"Report Number[s]*\s+([A-Za-z0-9]+)", value)
                if match:
                    return match.group(1).strip()
                # Check for standalone I-number format
                elif value.strip().startswith('I'):
                    if re.match(r'^I\d+T$', value.strip()):
                        return value.strip()
    return None

# Helper function to convert image to base64
def convert_image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")

# Helper function to create paths and URLs for images
def create_image_paths(image_filename, report_images_folder, base_url="http://example.com/Extracted_Images/Images"):
    abs_path = os.path.abspath(os.path.join(report_images_folder, image_filename))
    rel_path = os.path.relpath(abs_path, start=output_path)
    abs_url = f"{base_url}/{image_filename}"
    rel_url = f"/{rel_path.replace(os.sep, '/')}"
    
    return {"rel_path": rel_path, "rel_url": rel_url, "abs_path": abs_path, "abs_url": abs_url}

# Function to convert DataFrame to JSON format
import pandas as pd
import re

def df_to_key_value_json(df, image_data, report_number):
    json_data = {
        "report_number": report_number,
        "date": "",
        "stones": [],
        "assets": image_data  # This will contain the image paths and URLs
    }
    current_item = {}  # Dictionary to accumulate key-value pairs for each stone
    misc_keys = []  # List to accumulate keys with empty/null values
    date_pattern = re.compile(r"\b\d{2}\.\d{2}\.\d{4}\b")  # Pattern to match dates like dd.mm.yyyy
    report_number_pattern = re.compile(r"Report Number[s]*\s+([A-Za-z0-9]+)")  # Pattern for report numbers
    comment_pattern = re.compile(r"comments:(.*)", re.IGNORECASE)  # Pattern to capture comments

    # First, scan all columns for the date and top-level report number
    for col in df.columns:
        for idx, value in enumerate(df[col]):
            if isinstance(value, str):
                # Check for a date pattern
                if date_pattern.match(value):
                    json_data["date"] = value
                # Check for "Report Number" format for top-level report number
                report_number_match = report_number_pattern.search(value)
                if report_number_match and not json_data["report_number"]:
                    json_data["report_number"] = report_number_match.group(1).strip()

    # Process key-value pairs (columns A-B, C-D, etc.) as stone properties
    for col in range(0, len(df.columns), 2):
        if col + 1 < len(df.columns):  # Ensure we have a key-value pair
            key_col = df.columns[col]
            value_col = df.columns[col + 1]

            for idx, row in df.iterrows():
                key_value = row[key_col] if pd.notnull(row[key_col]) else None
                value_value = row[value_col] if pd.notnull(row[value_col]) else None

                if pd.isnull(key_value) and pd.isnull(value_value):
                    if current_item:  # Add accumulated data for a stone to stones array
                        json_data["stones"].append(current_item)
                        current_item = {}
                else:
                    if isinstance(key_value, str):
                        comment_match = comment_pattern.search(key_value)
                        if comment_match:
                            current_item["comments"] = comment_match.group(1).strip()
                            continue

                    if pd.isnull(value_value) or value_value == "":
                        if key_value and not isinstance(key_value, float):
                            misc_keys.append(str(key_value).strip())  # Keep original string without underscore
                    else:
                        # Process valid key-value pairs, ignoring "report number" for individual stones
                        if key_value:
                            key = str(key_value).strip().lower().replace(" ", "_")
                            if key != "report_number":  # Skip "report_number" for individual stones
                                current_item[key] = value_value

    # Append the last stone's details if any
    if current_item:
        json_data["stones"].append(current_item)

    # Add miscellaneous keys as a separate entry if needed
    if misc_keys:
        json_data["miscellaneous"] = misc_keys

    return json_data




# Function to process each Excel file
def process_excel_file(file_path, comparison_image_hashes):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    data = list(ws.values)
    df = pd.DataFrame(data)

    # Fix misaligned data in the first two columns
    for idx, row in df.iterrows():
        if pd.isnull(row[0]) and pd.notnull(row[1]):
            df.at[idx, 0] = df.at[idx, 1]
            df.at[idx, 1] = None

    df = df.dropna(axis=1, how='all')
    df = extract_report_numbers(df)

    # Get report number early for file naming
    report_number = get_report_number_from_df(df)
    
    # Generate base filename for output
    base_filename = f"Report_Number_{report_number}" if report_number else f"output_cleaned_{os.path.basename(file_path)}"
    output_excel_path = os.path.join(output_path, f'{base_filename}.xlsx')
    
    df.to_excel(output_excel_path, index=False, header=False)
    
    wb_output = load_workbook(output_excel_path)
    ws_output = wb_output.active
    last_column = ws_output.max_column + 1
    ws_output.column_dimensions[chr(64 + last_column)].width = 25

    # Create a report-specific subfolder for images
    if report_number:
        report_images_folder = os.path.join(images_output_path, report_number)
        if not os.path.exists(report_images_folder):
            os.makedirs(report_images_folder)
    else:
        report_images_folder = images_output_path

    image_data_paths = []  # For main JSON file (without base64)
    image_data_base64 = []  # For separate JSON file (with base64)

    for idx, image in enumerate(ws._images):
        img_filename = f"{report_number}_image_{idx+1}.png" if report_number else f"image_{idx+1}.png"
        img_path = os.path.join(report_images_folder, img_filename)

        # Save the image
        with open(img_path, 'wb') as img_file:
            img_file.write(image._data())

        # Get image paths and URLs
        path_info = create_image_paths(img_filename, report_images_folder)

        # Determine the image status
        if contains_barcode(img_path):
            status = "skipped "
        elif is_similar_image(img_path, comparison_image_hashes):
            status = "skipped "
        else:
            status = "saved"
            # Insert the image in Excel only if it is saved
            img = XLImage(img_path)
            img_cell = f'{chr(64 + last_column)}{idx + 2}'
            ws_output.add_image(img, img_cell)

        # Append path info for main JSON
        path_info["status"] = status
        image_data_paths.append(path_info)

        # For the separate JSON file, add base64 data regardless of status
        image_base64 = convert_image_to_base64(img_path)
        image_data_base64.append({"base64": image_base64, "status": status})

        # Remove skipped images from folder
        if status != "saved":
            os.remove(img_path)

    wb_output.save(output_excel_path)
    
    # Save main JSON file with paths
    json_output_path = os.path.join(output_path, f"{base_filename}.json")
    df_json_data = df_to_key_value_json(df, image_data_paths,report_number)

    with open(json_output_path, 'w') as json_file:
        json.dump(df_json_data, json_file, indent=4)

    # Save separate JSON with only images in base64 format
    images_json_output_path = os.path.join(output_path, f"{report_number}_images.json")
    with open(images_json_output_path, 'w') as img_json_file:
        json.dump({"images": image_data_base64}, img_json_file, indent=4)

    print(f"Processed file saved as {base_filename}")
    print(f"Images saved to {report_images_folder}")
    print(f"Main JSON data saved to {json_output_path}")
    print(f"Images JSON data saved to {images_json_output_path}")

# Load comparison image hashes
comparison_image_hashes = load_comparison_image_hashes(comparison_images_folder_path)

# Main execution loop
if __name__ == "__main__":
    # Iterate over each Excel file in the folder and process
    for file_name in os.listdir(main_folder_path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(main_folder_path, file_name)
            print(f"Processing file: {file_name}")
            process_excel_file(file_path, comparison_image_hashes)
